import flet as ft
import pandas as pd
import psycopg2
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

def main(page: ft.Page):
    page.title = "PostgreSQL View to Excel Exporter"
    page.theme_mode = ft.ThemeMode.LIGHT
    page.vertical_alignment = ft.MainAxisAlignment.START
    page.horizontal_alignment = ft.CrossAxisAlignment.CENTER
    page.padding = 20
    page.scroll = ft.ScrollMode.AUTO

    # Переменные для хранения значений
    db_host = ft.Ref[ft.TextField]()
    db_port = ft.Ref[ft.TextField]()
    db_name = ft.Ref[ft.TextField]()
    db_user = ft.Ref[ft.TextField]()
    db_password = ft.Ref[ft.TextField]()
    view_name = ft.Ref[ft.TextField]()
    template_path = ft.Ref[ft.TextField]()
    output_path = ft.Ref[ft.TextField]()
    status_text = ft.Ref[ft.Text]()

    def pick_template_result(e: ft.FilePickerResultEvent):
        if e.files:
            template_path.current.value = e.files[0].path
            template_path.current.update()

    def pick_output_result(e: ft.FilePickerResultEvent):
        if e.files:
            output_path.current.value = e.files[0].path
            output_path.current.update()
        elif e.path:
            output_path.current.value = e.path
            output_path.current.update()

    file_picker_template = ft.FilePicker(on_result=pick_template_result)
    file_picker_output = ft.FilePicker(on_result=pick_output_result)
    page.overlay.extend([file_picker_template, file_picker_output])

    def export_data(e):
        try:
            # Получение значений из полей ввода
            host = db_host.current.value
            port = db_port.current.value or "5432"
            database = db_name.current.value
            username = db_user.current.value
            password = db_password.current.value
            view = view_name.current.value
            template = template_path.current.value
            output = output_path.current.value or f"output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            # Валидация
            if not all([host, database, username, password, view, template]):
                status_text.current.value = "❌ Заполните все обязательные поля!"
                status_text.current.color = ft.Colors.RED
                status_text.current.update()
                return

            if not os.path.exists(template):
                status_text.current.value = "❌ Файл шаблона не найден!"
                status_text.current.color = ft.Colors.RED
                status_text.current.update()
                return

            # Обновление статуса
            status_text.current.value = "⏳ Подключаемся к базе данных..."
            status_text.current.color = ft.Colors.ORANGE
            status_text.current.update()

            # Подключение к PostgreSQL
            conn = psycopg2.connect(
                host=host,
                port=port,
                database=database,
                user=username,
                password=password
            )

            # Получение данных из view
            status_text.current.value = "⏳ Загружаем данные из view..."
            status_text.current.update()

            query = f"SELECT * FROM {view};"
            df = pd.read_sql_query(query, conn)
            conn.close()

            if df.empty:
                status_text.current.value = "⚠️ View не содержит данных!"
                status_text.current.color = ft.Colors.ORANGE
                status_text.current.update()
                return

            # Загрузка шаблона
            status_text.current.value = "⏳ Загружаем шаблон..."
            status_text.current.update()

            wb = load_workbook(template)
            ws = wb.active  # Используем активный лист

            # Очистка старых данных (со 2-й строки)
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)

            # Вставка данных
            status_text.current.value = "⏳ Вставляем данные в шаблон..."
            status_text.current.update()

            start_row = 2  # Начальная строка для данных
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start_row):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Сохранение
            status_text.current.value = "⏳ Сохраняем файл..."
            status_text.current.update()

            wb.save(output)

            # Успешное завершение
            status_text.current.value = f"✅ Данные успешно экспортированы!\nФайл: {output}\nЗаписей: {len(df)}"
            status_text.current.color = ft.Colors.GREEN
            status_text.current.update()

        except psycopg2.Error as e:
            status_text.current.value = f"❌ Ошибка базы данных: {e}"
            status_text.current.color = ft.Colors.RED
            status_text.current.update()
        except Exception as e:
            status_text.current.value = f"❌ Ошибка: {e}"
            status_text.current.color = ft.Colors.RED
            status_text.current.update()

    # Создание интерфейса
    page.add(
        ft.Text("Экспорт данных из PostgreSQL View в Excel", 
               size=24, weight=ft.FontWeight.BOLD),
        
        ft.Divider(),
        
        ft.Text("Настройки подключения к PostgreSQL", size=18, weight=ft.FontWeight.BOLD),
        
        ft.Row([
            ft.TextField(ref=db_host, label="Хост*", width=200, value="localhost"),
            ft.TextField(ref=db_port, label="Порт", width=100, value="5432"),
        ]),
        
        ft.Row([
            ft.TextField(ref=db_name, label="База данных*", width=200),
            ft.TextField(ref=db_user, label="Пользователь*", width=200),
        ]),
        
        ft.TextField(ref=db_password, label="Пароль*", password=True, can_reveal_password=True, width=400),
        
        ft.Divider(),
        
        ft.Text("Параметры экспорта", size=18, weight=ft.FontWeight.BOLD),
        
        ft.TextField(ref=view_name, label="Имя View*", width=400, hint_text="например, sales_report"),
        
        ft.Row([
            ft.TextField(ref=template_path, label="Путь к шаблону*", width=300, read_only=True),
            ft.ElevatedButton(
                "Выбрать шаблон",
                icon=ft.Icons.FILE_OPEN,
                on_click=lambda _: file_picker_template.pick_files(
                    allowed_extensions=["xlsx", "xls"],
                    dialog_title="Выберите файл шаблона Excel"
                )
            )
        ]),
        
        ft.Row([
            ft.TextField(ref=output_path, label="Путь для сохранения", width=300, 
                        hint_text="оставьте пустым для автоимени"),
            ft.ElevatedButton(
                "Выбрать папку",
                icon=ft.Icons.FOLDER_OPEN,
                on_click=lambda _: file_picker_output.get_directory_path(
                    dialog_title="Выберите папку для сохранения"
                )
            )
        ]),
        
        ft.Divider(),
        
        ft.ElevatedButton(
            "🚀 Экспортировать данные",
            on_click=export_data,
            style=ft.ButtonStyle(
                bgcolor=ft.Colors.BLUE,
                color=ft.Colors.WHITE,
                padding=20
            ),
            width=400
        ),
        
        ft.Divider(),
        
        ft.Text(ref=status_text, size=16, selectable=True, text_align=ft.TextAlign.CENTER),
        
        ft.Text("Обязательные поля помечены *", size=12, color=ft.Colors.GREY)
    )

if __name__ == "__main__":
    ft.app(target=main)