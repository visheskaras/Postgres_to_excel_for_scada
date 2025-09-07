import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from typing import Optional, Dict, Any, List
import os
from datetime import datetime
import logging
from dataclasses import dataclass

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class ExcelExportConfig:
    """Конфигурация экспорта в Excel"""
    template_path: str
    output_path: str
    sheet_name: str = "Data"
    start_row: int = 2
    start_col: int = 1
    preserve_formatting: bool = True
    auto_adjust_columns: bool = True

class ExcelExporter:
    """Класс для экспорта данных в Excel с использованием шаблонов"""
    
    def __init__(self, config: ExcelExportConfig):
        self.config = config
        self.workbook = None
        self.worksheet = None
    
    def load_template(self) -> bool:
        """Загрузка Excel шаблона"""
        try:
            if not os.path.exists(self.config.template_path):
                logger.error(f"Файл шаблона не найден: {self.config.template_path}")
                return False
            
            self.workbook = load_workbook(self.config.template_path)
            
            # Получение или создание листа
            if self.config.sheet_name in self.workbook.sheetnames:
                self.worksheet = self.workbook[self.config.sheet_name]
            else:
                logger.warning(f"Лист '{self.config.sheet_name}' не найден, используется активный лист")
                self.worksheet = self.workbook.active
                self.worksheet.title = self.config.sheet_name
            
            logger.info(f"Шаблон загружен: {self.config.template_path}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка загрузки шаблона: {e}")
            return False
    
    def clear_existing_data(self, start_row: Optional[int] = None, 
                          end_row: Optional[int] = None) -> bool:
        """
        Очистка существующих данных на листе
        
        Args:
            start_row: Начальная строка для очистки
            end_row: Конечная строка для очистки
        """
        if not self.worksheet:
            return False
        
        start_row = start_row or self.config.start_row
        end_row = end_row or self.worksheet.max_row
        
        try:
            if end_row >= start_row:
                self.worksheet.delete_rows(start_row, end_row - start_row + 1)
                logger.info(f"Очищены строки с {start_row} по {end_row}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка очистки данных: {e}")
            return False
    
    def export_data(self, df: pd.DataFrame, 
                   include_headers: bool = False) -> bool:
        """
        Экспорт данных DataFrame в Excel
        
        Args:
            df: DataFrame с данными
            include_headers: Включать ли заголовки колонок
        """
        if not self.worksheet or df.empty:
            return False
        
        try:
            # Преобразование DataFrame в строки
            rows = dataframe_to_rows(df, index=False, header=include_headers)
            
            # Вставка данных
            for row_idx, row in enumerate(rows, self.config.start_row):
                for col_idx, value in enumerate(row, self.config.start_col):
                    cell = self.worksheet.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Применение базового форматирования для новых данных
                    if include_headers and row_idx == self.config.start_row:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D3D3D3", 
                                              end_color="D3D3D3", 
                                              fill_type="solid")
            
            logger.info(f"Данные экспортированы: {len(df)} строк")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка экспорта данных: {e}")
            return False
    
    def auto_adjust_columns_width(self):
        """Автоматическая регулировка ширины колонок"""
        if not self.worksheet:
            return
        
        try:
            for column in self.worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Максимальная ширина 50
                self.worksheet.column_dimensions[column_letter].width = adjusted_width
            
            logger.info("Ширина колонок автоматически отрегулирована")
            
        except Exception as e:
            logger.warning(f"Не удалось отрегулировать ширину колонок: {e}")
    
    def save(self) -> bool:
        """Сохранение файла"""
        try:
            # Авторегулировка ширины колонок если включено
            if self.config.auto_adjust_columns:
                self.auto_adjust_columns_width()
            
            # Создание директории если не существует
            os.makedirs(os.path.dirname(self.config.output_path), exist_ok=True)
            
            self.workbook.save(self.config.output_path)
            logger.info(f"Файл сохранен: {self.config.output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка сохранения файла: {e}")
            return False
    
    def export_dataframe_to_template(self, df: pd.DataFrame, 
                                   clear_existing: bool = True,
                                   include_headers: bool = False) -> Dict[str, Any]:
        """
        Полный процесс экспорта DataFrame в шаблон
        
        Args:
            df: DataFrame для экспорта
            clear_existing: Очищать ли существующие данные
            include_headers: Включать ли заголовки колонок
            
        Returns:
            Словарь с результатом операции
        """
        try:
            # Загрузка шаблона
            if not self.load_template():
                return {"success": False, "message": "Ошибка загрузки шаблона"}
            
            # Очистка существующих данных
            if clear_existing:
                self.clear_existing_data()
            
            # Экспорт данных
            if not self.export_data(df, include_headers):
                return {"success": False, "message": "Ошибка экспорта данных"}
            
            # Сохранение
            if not self.save():
                return {"success": False, "message": "Ошибка сохранения файла"}
            
            return {
                "success": True,
                "message": "Данные успешно экспортированы",
                "output_path": self.config.output_path,
                "records_count": len(df)
            }
            
        except Exception as e:
            return {"success": False, "message": f"Ошибка экспорта: {e}"}
    
    def __enter__(self):
        """Контекстный менеджер"""
        self.load_template()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Контекстный менеджер - автоматическое сохранение"""
        if self.workbook:
            self.save()